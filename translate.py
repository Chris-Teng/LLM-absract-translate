from openpyxl import load_workbook
import requests
from requests.adapters import HTTPAdapter, Retry

Workbook = load_workbook(filename='output.xlsx')
sheet = Workbook.active

for i in range(1, sheet.max_row + 1):
    text = sheet.cell(row=i, column=2).value
    if text == None:
        continue

    s = requests.Session()
    retries = Retry(total=10,
                    backoff_factor=0.1,
                    status_forcelist=[ 500, 502, 503, 504 ])

    s.mount('http://', HTTPAdapter(max_retries=retries))

    url = "http://127.0.0.1:5000/v1/completions"
    headers = {
        'Content-Type': 'application/json',
    }
    data = {
        "prompt": f"Translate the text delimited by triple backticks to Chinese.```{text}```",
        "max_tokens": 4096,   # 一个token约等于4个单词
        "temperature": 1,
        "top_p": 0.9,
        "seed": 10
    }
    try:
        response = s.post(url, headers=headers, json=data, timeout=360)
        absText_CHN = response.json()['choices'][0]['text']
    except:
        try:
            response = s.post(url, headers=headers, json=data, timeout=360)
            absText_CHN = response.json()['choices'][0]['text']
        except:
            absText_CHN = 'fail'
    try:
        sheet.cell(row=i, column=3).value = absText_CHN
        Workbook.save('output_translate.xlsx')
    except:
        continue